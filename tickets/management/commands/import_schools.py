import os
import sys
import django
from django.core.management.base import BaseCommand, CommandError

# --- DISTRICT MAP for Secondary Schools ---
# The xlsx stores district as a number (1, 2, 3...). Map to readable names here.
SECONDARY_DISTRICT_MAP = {
    1: "District 1",
    2: "District 2",
    3: "District 3",
    4: "District 4",
    5: "District 5",
    6: "District 6",
    7: "District 7",
    8: "District 8",
}


class Command(BaseCommand):
    help = 'Imports school data directly from a .xlsx (Excel) file.'

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='The path to the schools .xlsx file (e.g. schools.xlsx).'
        )

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        added_count = 0
        skipped_count = 0

        # --- 1. Check that openpyxl is installed ---
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl is not installed. Run: pip install openpyxl"
            )

        # --- 2. Check the file exists and is an xlsx ---
        if not os.path.isfile(file_path):
            raise CommandError(
                f"File not found: '{file_path}'. "
                f"Make sure you are running the command from the same folder as the file, "
                f"or provide the full path."
            )

        if not file_path.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            raise CommandError(
                f"'{file_path}' does not appear to be an Excel file. "
                f"Please export your data as .xlsx and try again."
            )

        # --- 3. Load the workbook ---
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Could not open workbook: {e}")

        ws = wb.active  # Uses the first/active sheet

        self.stdout.write(
            self.style.NOTICE(
                f"Reading sheet '{ws.title}' ({ws.max_row - 1} data rows)..."
            )
        )

        from tickets.models import School

        # --- 4. Iterate rows (skip header row 1) ---
        for row in ws.iter_rows(min_row=2, values_only=True):
            # ---- LEFT SIDE: Secondary Schools ----
            # Col A (idx 0): District number | Col C (idx 2): School ID | Col D (idx 3): School Name
            sec_district_num = row[0]
            sec_id = row[2]
            sec_name = row[3]

            if sec_id and str(sec_id).strip() and sec_name and str(sec_name).strip():
                district_label = SECONDARY_DISTRICT_MAP.get(
                    int(sec_district_num),
                    f"District {sec_district_num}"
                ) if sec_district_num else "Unknown"

                _, created = School.objects.get_or_create(
                    school_id=str(int(sec_id)).strip(),
                    defaults={
                        'name': str(sec_name).strip(),
                        'district': district_label,
                        'password': 'pbkdf2_sha256$1200000$yxVstmeDa4jgFPKbARyyYt$TYZVP2LaT6q3jtk6FzUV75dzoGUi4DD+7LfmWAAsrlE='
                    }
                )
                if created:
                    added_count += 1
                else:
                    skipped_count += 1

            # ---- RIGHT SIDE: Elementary Schools ----
            # Col F (idx 5): District name | Col H (idx 7): School ID | Col I (idx 8): School Name
            elem_district = row[5]
            elem_id = row[7]
            elem_name = row[8]

            if elem_id and str(elem_id).strip() and elem_name and str(elem_name).strip():
                _, created = School.objects.get_or_create(
                    school_id=str(int(elem_id)).strip(),
                    defaults={
                        'name': str(elem_name).strip(),
                        'district': str(elem_district).strip() if elem_district else "Unknown",
                        'password': 'pbkdf2_sha256$1200000$yxVstmeDa4jgFPKbARyyYt$TYZVP2LaT6q3jtk6FzUV75dzoGUi4DD+7LfmWAAsrlE='
                    }
                )
                if created:
                    added_count += 1
                else:
                    skipped_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"\n✅ Done! {added_count} school(s) imported, {skipped_count} already existed (skipped)."
            )
        )